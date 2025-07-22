import os
import copy
from CargadorCarrera import CargadorCarrera
from modelos.Carrera import Carrera

from openpyxl import Workbook
from openpyxl.styles import Alignment

def limpiarPantalla():
    os.system("cls" if os.name == "nt" else "clear")

def mostrarMenu(carreraSeleccionada: Carrera, asignaturasSeleccionadas):
    limpiarPantalla()
    seleccionadas = ', '.join([a.getSigla() for a in asignaturasSeleccionadas]) if asignaturasSeleccionadas else "Ninguna"
    print(f"""Hola, ¿qué deseas hacer?
1. Seleccionar carrera
2. Ver asignaturas (Carrera seleccionada: {carreraSeleccionada.getNombre()})
3. Seleccionar asignaturas (Actualmente seleccionadas: {seleccionadas})
4. Ver horario
5. Ver datos
6. Exportar horario a Excel
7. Salir
""")

def seleccionarCarrera(carreras: list) -> Carrera:
    limpiarPantalla()
    print("Lista de carreras disponibles:\n")
    for i, carrera in enumerate(carreras):
        print(f"{i + 1}. [{carrera.getNombre()}]")

    try:
        opcion = int(input("\nSelecciona una carrera (número): "))
        if 1 <= opcion <= len(carreras):
            return carreras[opcion - 1]
        else:
            print("Opción inválida.")
    except ValueError:
        print("Entrada no válida.")
    
    input("\nPresiona Enter para continuar...")
    return None

def mostrarAsignaturas(carrera: Carrera):
    limpiarPantalla()
    if carrera.getNombre() == "Ninguna":
        print("Debes seleccionar una carrera primero.")
    else:
        print(f"Asignaturas de la carrera {carrera.getNombre()}:\n")
    print(f"{'N°':<3} {'Sigla':<10} {'Nombre':<45} {'Plan':<10} {'Nivel':<10} {'# Secciones':<12}")
    print("-" * 95)

    asignaturas = carrera.getAsignaturas()

    # Ordenar por plan, luego por nivel
    asignaturas.sort(key=lambda a: (str(a.getPlan()), str(a.getNivel())))

    for i, asignatura in enumerate(asignaturas, start=1):
        sigla = asignatura.getSigla()
        nombre = asignatura.getNombre()[:44]  # Evita que se desborde
        plan = asignatura.getPlan()
        nivel = asignatura.getNivel()
        cantidad = len(asignatura.getSeccion())
        print(f"{i:<3} {sigla:<10} {nombre:<45} {plan:<10} {nivel:<10} {cantidad:<12}")

    input("\nPresiona Enter para continuar...")

def horariosChocan(horariosOcupados, nuevosHorarios):
    for nuevo in nuevosHorarios:
        for ocupado in horariosOcupados:
            if nuevo == ocupado["horario"]:
                return ocupado  # Retorna la info de la asignatura y sección que choca
    return None


def seleccionarAsignaturas(carrera: Carrera):
    limpiarPantalla()
    asignaturasSeleccionadas = []
    horariosOcupados = []

    if carrera.getNombre() == "Ninguna":
        print("Debes seleccionar una carrera primero.")
        input("\nPresiona Enter para continuar...")
        return asignaturasSeleccionadas

    print("Selecciona las asignaturas (ingresa los números separados por coma, por ejemplo: 1,3,4):\n")
    print(f"\nAsignaturas de la carrera {carrera.getNombre()}:\n")
    asignaturas = carrera.getAsignaturas()

    # Ordenar por plan, luego por nivel
    asignaturas.sort(key=lambda a: (str(a.getPlan()), str(a.getNivel())))

    print(f"{'N°':<3} {'Sigla':<10} {'Nombre':<45} {'Plan':<10} {'Nivel':<10} {'# Secciones':<12}")
    print("-" * 95)

    for i, asignatura in enumerate(asignaturas, start=1):
        sigla = asignatura.getSigla()
        nombre = asignatura.getNombre()[:44]  # Evita que se desborde
        plan = asignatura.getPlan()
        nivel = asignatura.getNivel()
        cantidad = len(asignatura.getSeccion())
        print(f"{i:<3} {sigla:<10} {nombre:<45} {plan:<10} {nivel:<10} {cantidad:<12}")

    entrada = input("\nSelecciona: ")
    try:
        indices = [int(i.strip()) - 1 for i in entrada.split(",")]

        for idx in indices:
            if 0 <= idx < len(asignaturas):
                asignaturaOriginal = asignaturas[idx]
                secciones = asignaturaOriginal.getSeccion()

                if not secciones:
                    print(f"La asignatura {asignaturaOriginal.getSigla()} no tiene secciones registradas.")
                    continue

                seleccion_valida = False
                while not seleccion_valida:
                    limpiarPantalla()
                    print(f"\nSecciones disponibles para [{asignaturaOriginal.getSigla()}] - [{asignaturaOriginal.getNombre()}]:\n")

                    for j, seccion in enumerate(secciones, start=1):
                        print(f"{j}. Sección: [{seccion.getCodigo()}] | Docente: [{seccion.getDocente()}]")
                        for horario in seccion.getHorario():
                            print(f"[{horario}]")
                        print()  # Espacio entre secciones

                    seleccion = input("Selecciona una sección (número, o 0 para cancelar): ")

                    try:
                        idx_seccion = int(seleccion.strip()) - 1

                        if idx_seccion == -1:
                            print("Selección de sección cancelada.")
                            break

                        if 0 <= idx_seccion < len(secciones):
                            seccionElegida = secciones[idx_seccion]
                            conflicto = horariosChocan(horariosOcupados, seccionElegida.getHorario())

                            if conflicto:
                                print("\n¡La sección seleccionada tiene un tope horario!")
                                print(f"Conflicto con: [{conflicto['sigla']}] - [{conflicto['nombre']}] Sección: [{conflicto['seccion']}]")
                                print(f"Horario en conflicto: [{conflicto['horario']}]\n")
                                print("Intenta con otra sección o ingresa 0 para omitir esta asignatura.\n")
                                input("Presiona Enter para continuar...")
                            else:
                                asignaturaClonada = copy.deepcopy(asignaturaOriginal)
                                asignaturaClonada.setSeccion([seccionElegida])
                                asignaturasSeleccionadas.append(asignaturaClonada)

                                for horario in seccionElegida.getHorario():
                                    horariosOcupados.append({
                                        "horario": horario,
                                        "sigla": asignaturaOriginal.getSigla(),
                                        "nombre": asignaturaOriginal.getNombre(),
                                        "seccion": seccionElegida.getCodigo()
                                    })
                                seleccion_valida = True
                        else:
                            print("Número de sección inválido.")
                            input("Presiona Enter para continuar...")
                    except ValueError:
                        print("Entrada inválida.")
                        input("Presiona Enter para continuar...")
    except ValueError:
        print("Entrada no válida.")

    input("\nPresiona Enter para continuar...")
    return asignaturasSeleccionadas


def verHorario(asignaturasSeleccionadas):
    limpiarPantalla()
    if not asignaturasSeleccionadas:
        print("No has seleccionado asignaturas aún.")
    else:
        print("Horario generado a partir de las asignaturas seleccionadas:\n")

        for asignatura in asignaturasSeleccionadas:
            print(f"Sigla: [{asignatura.getSigla()}] - [{asignatura.getNombre()}]")
            print(f"Plan: [{asignatura.getPlan()}]\n")

            for seccion in asignatura.getSeccion():
                print(f"Sección: [{seccion.getCodigo()}] | Docente: [{seccion.getDocente()}]")
                for horario in seccion.getHorario():
                    print(f"[{horario}]")
            print("-" * 80)
    input("Presiona Enter para continuar...")

def generarExcelHorario(asignaturasSeleccionadas):
    if not asignaturasSeleccionadas:
        print("No hay asignaturas seleccionadas para generar el horario.")
        input("Presiona Enter para continuar...")
        return

    dias_abreviados = {
        "Lu": "Lunes",
        "Ma": "Martes",
        "Mi": "Miércoles",
        "Ju": "Jueves",
        "Vi": "Viernes",
        "Sa": "Sábado",
        "Do": "Domingo"
    }

    dias_semana = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]
    bloques = [
        "08:31-09:50", "10:01-10:40", "10:41-11:20", "11:31-12:10",
        "13:41-14:20", "14:31-15:10", "16:15-17:45", "18:00-19:30",
        "19:01-20:20", "20:31-21:10", "21:11-22:30"
    ]

    # Diccionario que usaremos para llenar la tabla
    horario = {bloque: {dia: "" for dia in dias_semana} for bloque in bloques}

    for asignatura in asignaturasSeleccionadas:
        for seccion in asignatura.getSeccion():
            for horario_raw in seccion.getHorario():
                try:
                    partes = horario_raw.strip().split(" ")
                    if len(partes) != 4 or partes[2] != "-":
                        continue  # formato incorrecto

                    dia_abrev = partes[0]
                    hora_inicio = partes[1][:5]  # "11:31:00" -> "11:31"
                    hora_fin = partes[3][:5]     # "12:10:00" -> "12:10"
                    bloque = f"{hora_inicio}-{hora_fin}"

                    dia = dias_abreviados.get(dia_abrev)

                    print(f"[DEBUG] Día: {dia}, Bloque: {bloque}")

                    if dia and bloque in horario and dia in horario[bloque]:
                        contenido = f"{seccion.getCodigo()}"
                        if horario[bloque][dia]:
                            horario[bloque][dia] += " / " + contenido
                        else:
                            horario[bloque][dia] = contenido
                except Exception as e:
                    print(f"[ERROR] {e}")
                    continue

    # Crear archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Horario"

    # Escribir encabezados
    ws.cell(row=1, column=1, value="Bloque / Día")
    for col, dia in enumerate(dias_semana, start=2):
        ws.cell(row=1, column=col, value=dia)

    # Escribir contenido
    for row, bloque in enumerate(bloques, start=2):
        ws.cell(row=row, column=1, value=bloque)
        for col, dia in enumerate(dias_semana, start=2):
            celda = ws.cell(row=row, column=col, value=horario[bloque][dia])
            celda.alignment = Alignment(wrap_text=True, vertical="top")

    archivo_nombre = "horario_generado.xlsx"
    wb.save(archivo_nombre)
    print(f"\nHorario guardado exitosamente en '{archivo_nombre}'.")
    input("Presiona Enter para continuar...")


def verDatos(listaCarrera):
    limpiarPantalla()

    for carrera in listaCarrera:
        carrera.imprimirCarrera()

    input("Presiona Enter para continuar...")

def main():
    carreras = CargadorCarrera.cargarCarreras()
    carreraSeleccionada = Carrera()
    carreraSeleccionada.setNombre("Ninguna")
    asignaturasSeleccionadas = []

    while True:
        mostrarMenu(carreraSeleccionada, asignaturasSeleccionadas)
        opcion = input("Selecciona una opción: ")

        if opcion == "1":
            seleccion = seleccionarCarrera(carreras)
            if seleccion:
                carreraSeleccionada = seleccion
                asignaturasSeleccionadas = []  # reset al cambiar carrera
        elif opcion == "2":
            mostrarAsignaturas(carreraSeleccionada)
        elif opcion == "3":
            asignaturasSeleccionadas = seleccionarAsignaturas(carreraSeleccionada)
        elif opcion == "4":
            verHorario(asignaturasSeleccionadas)
        elif opcion == "5":
            verDatos(carreras)
        elif opcion == "6":
            generarExcelHorario(asignaturasSeleccionadas)
        elif opcion == "7":
            print("Saliendo...")
            break
        else:
            print("Opción no válida.")
            input("Presiona Enter para continuar...")

if __name__ == "__main__":
    main()
